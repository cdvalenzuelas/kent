import pandas as pd
from openpyxl import Workbook, load_workbook
import re

# Funciones de cada tipo de válvulas
from src.modules.hd.ball import ball
from src.modules.hd.gate import gate
from src.modules.hd.check import check
from src.modules.hd.monoflange import monoflange
from src.modules.hd.globe import globe

# Crear un índice común para las válvulas


def valves_index(row):
    spec, size, tag = row

    return f'{spec} {size} {tag}'


def hd(mto_df):
    # Hacer una copia el mto_df
    mto_df = mto_df.copy()

    # Extrayendo únicamente las válvulas
    mto_df = mto_df[mto_df['TYPE'] == 'VL']

    # Si existen válvulas hacer las hojas de datos de válvulas
    if mto_df.shape[0] > 0:

        # leer el archivo de válvulas
        valves_df = pd.read_csv(
            './src/clients/cenit/elements/cenit_valves_tags.csv')

        # Leer los templates de válvulas
        wb = load_workbook(
            filename='./src/clients/cenit/templates/hd_template.xlsx')

        # Extrayendo Las columnas necesarias del MTO
        mto_df = mto_df[['SPEC', 'FIRST_SIZE_NUMBER', 'TAG']]

        # Eliminando válvulas repetidas del MTO
        mto_df = mto_df.drop_duplicates(
            ['SPEC', 'FIRST_SIZE_NUMBER', 'TAG'], keep='last')

        # Crear un índice común entre el mto y la tabla de válvulas
        valves_df['valves_index'] = valves_df[['SPEC', 'SIZE_NUMBER', 'TAG']].apply(
            valves_index, axis=1)

        mto_df['valves_index'] = mto_df[['SPEC', 'FIRST_SIZE_NUMBER', 'TAG']].apply(
            valves_index, axis=1)

        # Crear el merge entre las dos tablas
        valves_df = pd.merge(mto_df, valves_df, how='left', on='valves_index')

        valves_df.drop(['SPEC_y', 'TAG_y', 'valves_index',
                        'FIRST_SIZE_NUMBER'], inplace=True, axis=1)

        valves_df.fillna('-', inplace=True)

        valves_df.rename({'SPEC_x': 'SPEC', 'TAG_x': 'TAG'},
                         inplace=True, axis=1)

        # ORDENAR EL DATAFRAME
        valves_df.sort_values(by=['TAG', 'SPEC', 'SIZE_NUMBER'], inplace=True)

        # Resetear los índices
        valves_df.reset_index(inplace=True, drop=True)

        # Recoger los diámetros por cada TAG y por cada SPEC
        final_index = valves_df.shape[0] - 1

        valves_sizes = []

        size_per_tag_and_spec = ''

        # VER SOBR QUE COLUNMAS SE VAN A COMPARAR LAS VÁLVULAS

        valves_df_columns = [
            prop for prop in valves_df.columns if prop not in ['SIZE', 'SIZE_NUMBER']]

        for index, row in valves_df.iterrows():
            if index < final_index:
                # Extraer dos filas consecutivas
                row_a = valves_df.iloc[index]
                row_b = valves_df.iloc[index + 1]

                cond = True

                # Comparar si las dós filas consecutivas difieren únicamente en tamaño
                for prop in valves_df_columns:
                    cond = cond and (row_a[prop] == row_b[prop])

                    # Si la condición pasa a ser falsa se acaba la comparación
                    if not cond:
                        break

                # Si las dos válvulas consecutivas son iguales
                if cond:
                    size_per_tag_and_spec += f'{row_a["SIZE"]}, '

                    # Si las dos últimas válvulas son iguales
                    if index == final_index - 1:
                        size_per_tag_and_spec += f'{row_b["SIZE"]}'
                        size_per_tag_and_spec = size_per_tag_and_spec.replace(
                            '\"', '')
                        valves_sizes.append(size_per_tag_and_spec)
                        break
                # Si las dos válvulas consecutivas son diferentes
                else:
                    size_per_tag_and_spec += f'{row_a["SIZE"]}'
                    size_per_tag_and_spec = size_per_tag_and_spec.replace(
                        '\"', '')
                    valves_sizes.append(size_per_tag_and_spec)
                    size_per_tag_and_spec = ''

                    # Si las dos últimas válvulas son diferentes
                    if index == final_index - 1:
                        first_size_b = row_b["SIZE"].replace('\"', '')
                        valves_sizes.append(first_size_b)
                        break

        # Eliminar válvulas repetidas
        valves_df = valves_df.drop_duplicates(['SPEC', 'TAG', 'TYPE_CODE', 'RATING',
                                               'SERVICE', 'DESIGN_PRESSURE', 'DESIGN_TEMPERATURE', 'TYPE', 'CONECTION',
                                               'END_CODE', 'BORE', 'OPERATOR', 'BODY', 'BONNET_GASKET', 'BALL', 'STEM',
                                               'OTHER', 'SEALS', 'SEAT', 'PACKING', 'API_TRIM', 'LENGTH', 'DESIGN_CODE',
                                               'TEST_PRESSURE', 'APPLICABLE_STANDARDS', 'BORE_ACCORDING_TO', 'COATING',
                                               'SYSTEM', 'PATTERN'], keep='last')

        valves_df.reset_index(inplace=True, drop=True)

        # UNIR LOS VALVES_SIZES AL DATA_FRAME
        new_column = pd.DataFrame({'SIZES': valves_sizes})

        valves_df = pd.concat([new_column, valves_df], axis=1)

        # Iterar todo el dataframe
        for index, row in valves_df.iterrows():
            if row['TYPE_CODE'] in ['BAL', 'BAL6']:
                wb = ball(row, wb, valve_type='BALL')
            elif row['TYPE_CODE'] in ['GAT']:
                wb = gate(row, wb, valve_type='GATE')
            elif row['TYPE_CODE'] in ['CHL', 'CHS']:
                wb = check(row, wb, valve_type='CHECK')
            elif row['TYPE_CODE'] in ['DBB']:
                wb = monoflange(row, wb, valve_type='MONOFLANGE')
            elif row['TYPE_CODE'] in ['GLB']:
                wb = globe(row, wb, valve_type='GLOBE')
            else:
                print(
                    f"❌ NO SE HAN CREADO LAS HOJAS DE DATOS DE {row['TAG']}\n")

        # Eliminar hojas dummy
        std = wb.get_sheet_by_name('BALL')
        wb.remove_sheet(std)

        std = wb.get_sheet_by_name('GATE')
        wb.remove_sheet(std)

        std = wb.get_sheet_by_name('GLOBE')
        wb.remove_sheet(std)

        std = wb.get_sheet_by_name('CHECK')
        wb.remove_sheet(std)

        std = wb.get_sheet_by_name('MONOFLANGE')
        wb.remove_sheet(std)

        # Si no se tienen hojas de datos (válvulas) que no se imprima o guarde el documento
        if not(len(wb.sheetnames) == 1 and wb.sheetnames[0] == "ESRI_MAPINFO_SHEET"):
            # Guardar el workbook
            wb.save('./output/hd.xlsx')
